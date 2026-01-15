#!/usr/bin/env python
"""Extended model comparison for the European Expansion case study."""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.compose import ColumnTransformer
from sklearn.ensemble import (
    ExtraTreesRegressor,
    GradientBoostingRegressor,
    HistGradientBoostingRegressor,
    RandomForestRegressor,
)
from sklearn.impute import SimpleImputer
from sklearn.metrics import (
    mean_absolute_error,
    mean_absolute_percentage_error,
    mean_squared_error,
    r2_score,
)
from sklearn.model_selection import KFold, cross_val_score, train_test_split
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import OneHotEncoder

from lightgbm import LGBMRegressor

try:
    from xgboost import XGBRegressor

    HAS_XGB = True
except ImportError:
    HAS_XGB = False


RANDOM_STATE = 42
DATA_PATH = Path("store_database_for_homework.xlsx")
HEADER_ROW = 2
BLUE_THEME_INDICES = [3, 4]
TARGET_COL = "Annual Total Gross Turnover*"
LATEST_YEAR = 2023
OUTPUT_PATH = Path("outputs/extended_model_comparison.csv")


def detect_blue_columns(path: Path, header_row: int, theme_indices: List[int]) -> List[str]:
    """Return workbook columns whose fill color theme matches the provided indices."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    columns: List[str] = []
    for cell in ws[header_row]:
        theme = getattr(cell.fill.start_color, "theme", None)
        if theme in theme_indices and cell.value:
            columns.append(cell.value)
    wb.close()
    return columns


def build_pipeline(
    numeric: List[str],
    categorical: List[str],
    estimator,
) -> Pipeline:
    """Create the preprocessing + estimator pipeline used for every model."""
    num_transformer = Pipeline([("imputer", SimpleImputer(strategy="median"))])
    cat_transformer = Pipeline(
        [
            ("imputer", SimpleImputer(strategy="most_frequent")),
            ("onehot", OneHotEncoder(handle_unknown="ignore")),
        ]
    )
    preprocessor = ColumnTransformer(
        [
            ("num", num_transformer, numeric),
            ("cat", cat_transformer, categorical),
        ]
    )
    return Pipeline([("preprocessor", preprocessor), ("model", estimator)])


def evaluate_pipeline(
    name: str,
    estimator,
    numeric_features: List[str],
    categorical_features: List[str],
    X_train: pd.DataFrame,
    y_train: pd.Series,
    X_test: pd.DataFrame,
    y_test: pd.Series,
    cv: KFold,
) -> Dict[str, float]:
    """Fit, score, and summarise a single pipeline."""
    pipeline = build_pipeline(numeric_features, categorical_features, estimator)
    pipeline.fit(X_train, y_train)
    preds = pipeline.predict(X_test)
    metrics = {
        "model": name,
        "rmse": float(np.sqrt(mean_squared_error(y_test, preds))),
        "mae": float(mean_absolute_error(y_test, preds)),
        "mape": float(mean_absolute_percentage_error(y_test, preds)),
        "r2": float(r2_score(y_test, preds)),
    }
    cv_scores = cross_val_score(
        pipeline,
        X_train,
        y_train,
        scoring="neg_root_mean_squared_error",
        cv=cv,
    )
    metrics["cv_rmse_mean"] = float(-cv_scores.mean())
    metrics["cv_rmse_std"] = float(cv_scores.std())
    return metrics


def main() -> None:
    raw_df = pd.read_excel(DATA_PATH, header=HEADER_ROW - 1)
    blue_columns = detect_blue_columns(DATA_PATH, HEADER_ROW, BLUE_THEME_INDICES)
    feature_cols = [col for col in blue_columns if col != TARGET_COL]

    df = raw_df[raw_df["Year"] == LATEST_YEAR].copy()
    df["Location type"] = df["Location type"].fillna("Unknown").astype(str).str.strip()
    X = df[feature_cols]
    y = df[TARGET_COL]

    numeric_features = [col for col in feature_cols if pd.api.types.is_numeric_dtype(df[col])]
    categorical_features = [col for col in feature_cols if col not in numeric_features]

    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=0.2, random_state=RANDOM_STATE
    )
    cv = KFold(n_splits=5, shuffle=True, random_state=RANDOM_STATE)

    estimators: Dict[str, object] = {
        "Random Forest": RandomForestRegressor(
            n_estimators=800,
            min_samples_leaf=3,
            random_state=RANDOM_STATE,
            n_jobs=-1,
        ),
        "HistGradientBoosting": HistGradientBoostingRegressor(
            max_iter=1000,
            learning_rate=0.05,
            min_samples_leaf=15,
            l2_regularization=0.1,
            random_state=RANDOM_STATE,
        ),
        "LightGBM": LGBMRegressor(
            n_estimators=1600,
            learning_rate=0.03,
            num_leaves=31,
            subsample=0.8,
            colsample_bytree=0.9,
            reg_lambda=0.8,
            random_state=RANDOM_STATE,
        ),
        "ExtraTrees": ExtraTreesRegressor(
            n_estimators=1200,
            min_samples_leaf=2,
            max_features="sqrt",
            random_state=RANDOM_STATE,
            n_jobs=-1,
        ),
        "GradientBoosting": GradientBoostingRegressor(
            n_estimators=2500,
            learning_rate=0.02,
            max_depth=3,
            subsample=0.9,
            random_state=RANDOM_STATE,
        ),
    }
    if HAS_XGB:
        estimators["XGBoost"] = XGBRegressor(
            n_estimators=2000,
            learning_rate=0.03,
            max_depth=6,
            subsample=0.8,
            colsample_bytree=0.8,
            reg_lambda=1.0,
            random_state=RANDOM_STATE,
            objective="reg:squarederror",
            n_jobs=-1,
        )

    results = [
        evaluate_pipeline(
            name,
            estimator,
            numeric_features,
            categorical_features,
            X_train,
            y_train,
            X_test,
            y_test,
            cv,
        )
        for name, estimator in estimators.items()
    ]

    comparison_df = pd.DataFrame(results).sort_values("rmse").reset_index(drop=True)
    OUTPUT_PATH.parent.mkdir(exist_ok=True, parents=True)
    comparison_df.to_csv(OUTPUT_PATH, index=False)
    print(comparison_df)
    best_row = comparison_df.iloc[0]
    print(
        f"\nBest model: {best_row['model']} "
        f"(RMSE={best_row['rmse']:.1f}, R2={best_row['r2']:.3f}, "
        f"CV RMSE={best_row['cv_rmse_mean']:.1f}±{best_row['cv_rmse_std']:.1f})"
    )


if __name__ == "__main__":
    main()
